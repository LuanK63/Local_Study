"""
research/generate_thesis_report.py
===================================
Sinh báo cáo đầy đủ cho khóa luận:
  1. Xuất Excel chi tiết (summary + per-question + chunk stats)
  2. Phân tích câu hỏi truy xuất sai (miss / partial) từ _details.json
  3. Vẽ biểu đồ so sánh 7 chiến lược chunking
"""
import sys, os, json
import sqlite3
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent.parent))

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not installed. Excel export will be skipped.")

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent.parent
DB_PATH     = BASE_DIR / "data" / "benchmark_logs.db"
DETAILS_DIR = BASE_DIR / "artifacts" / "reports" / "details"
OUT_DIR     = BASE_DIR / "artifacts" / "reports"
OUT_DIR.mkdir(parents=True, exist_ok=True)

import sys
DATASET_VERSION = "v5.0-dpr_containment"
for arg in sys.argv[1:]:
    if arg.startswith("--dataset-version="):
        DATASET_VERSION = arg.split("=")[1]

DATASET_LABEL   = f"{DATASET_VERSION} – 60 câu hỏi"
STRATEGY_ORDER = ["fixed", "recursive", "token", "sentence", "paragraph", "semantic", "parent_child"]
STRATEGY_LABELS = {
    "fixed":        "Fixed-Size",
    "recursive":    "Recursive",
    "token":        "Token",
    "sentence":     "Sentence",
    "paragraph":    "Paragraph",
    "semantic":     "Semantic",
    "parent_child": "Parent-Child",
}

# ─── DB helper ───────────────────────────────────────────────────────────────
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ─── 1. Fetch aggregate data ─────────────────────────────────────────────────
def fetch_summary():
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("""
        SELECT e.experiment_id, e.config_name,
               r.total_questions,
               r.precision_at_5, r.recall_at_5, r.f1_at_5,
               r.hit_rate_at_5, r.mrr, r.retrieval_latency,
               s.total_chunks, s.avg_chunk_size, s.median_chunk_size,
               s.min_chunk_size, s.max_chunk_size, s.ingestion_time_seconds
        FROM experiments e
        JOIN benchmark_results r ON e.experiment_id = r.experiment_id
        LEFT JOIN chunk_statistics s ON e.experiment_id = s.experiment_id
        WHERE e.dataset_version = ?
        ORDER BY r.mrr DESC
    """, (DATASET_VERSION,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

# ─── 2. Fetch per-question data from DB ──────────────────────────────────────
def fetch_per_question():
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("""
        SELECT e.config_name, q.question_id, q.hit_rate, q.mrr,
               q.retrieval_latency, q.ref_coverage, q.full_hit
        FROM question_results q
        JOIN experiments e ON q.experiment_id = e.experiment_id
        WHERE e.dataset_version = ?
        ORDER BY e.config_name, q.question_id
    """, (DATASET_VERSION,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

# ─── 3. Analyse _details.json ────────────────────────────────────────────────
def load_all_details():
    """Load tất cả file _details.json → dict {strategy: [question_detail, ...]}"""
    all_details = {}
    for strat in STRATEGY_ORDER:
        path = DETAILS_DIR / f"{strat}_details.json"
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                all_details[strat] = json.load(f)
        else:
            all_details[strat] = []
    return all_details

def analyse_errors(all_details):
    """ Phân tích câu hỏi bị miss (hit=0) và partial (hit>0 nhưng full_hit=0).
    Trả về:
      - miss_table  : list of dicts (cross-strategy miss)
      - per_strategy: dict {strat: {'miss': [...], 'partial': [], 'hit': []}}
    """
    per_strategy = {}

    for strat, questions in all_details.items():
        miss    = []
        partial = []
        hit     = []
        for q in questions:
            hr  = q.get("metrics", {}).get("hit_rate", 0)
            fh  = q.get("metrics", {}).get("full_hit", 0)
            qid = q.get("question_id", "")
            qtxt= q.get("question", "")
            gt_pages = q.get("ground_truth_pages", [])
            gt_docs  = q.get("ground_truth_docs", [])
            # retrieved pages top-5
            ret_pages = [c.get("page_num", 0) for c in q.get("retrieved_chunks", [])[:5]]
            ret_docs  = [c.get("doc_name", "") for c in q.get("retrieved_chunks", [])[:5]]

            entry = {
                "question_id": qid,
                "question": qtxt[:100],
                "gt_docs": gt_docs,
                "gt_pages": gt_pages,
                "ret_pages": ret_pages,
                "ret_docs": [os.path.basename(d) for d in ret_docs],
                "hit_rate": hr,
                "full_hit": fh,
                "mrr": q.get("metrics", {}).get("mrr", 0),
            }
            if hr == 0:
                miss.append(entry)
            elif fh == 0:
                partial.append(entry)
            else:
                hit.append(entry)

        per_strategy[strat] = {"miss": miss, "partial": partial, "hit": hit}

    # Cross-strategy: câu nào bị MISS ở nhiều strategy nhất
    from collections import Counter
    miss_counter = Counter()
    miss_info = {}
    for strat, d in per_strategy.items():
        for q in d["miss"]:
            qid = q["question_id"]
            miss_counter[qid] += 1
            miss_info[qid] = q  # lấy info từ lần cuối (giống nhau)

    miss_table = []
    for qid, cnt in miss_counter.most_common():
        info = miss_info[qid]
        miss_table.append({**info, "miss_count": cnt})

    return miss_table, per_strategy

# ─── 4. Excel Export ─────────────────────────────────────────────────────────
def make_header_style(ws, row, cols, fill_hex="1F4E79", font_color="FFFFFF"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=font_color, size=11)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def auto_col_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

ALT_FILLS = [PatternFill("solid", fgColor="EBF3FB"), PatternFill("solid", fgColor="FFFFFF")]

def export_excel(summary_rows, per_q_rows, per_strategy, miss_table, out_path=None):
    if not HAS_OPENPYXL:
        print("[SKIP] openpyxl not available.")
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: Tổng quan ───────────────────────────────────────────────────
    ws1 = wb.active
   
    ws1.title = "Tổng quan"
    ws1["A1"] = "BẢNG SO SÁNH CÁC CHIẾN LƯỢC CHUNKING – DATASET v5.0 (60 câu)"
    ws1["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws1.merge_cells("A1:O1")
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 6  # spacer

    headers = [
        "Chiến lược", "Số chunks", "Avg size (chars)", "Median size", "Min", "Max",
        "Build index (s)",
        "Precision@5", "Recall@5", "F1@5",
        "HitRate@5", "MRR", "Latency (ms)" ]
    ws1.append([])
    ws1.append(headers)
    make_header_style(ws1, ws1.max_row, len(headers))

    # Sort by MRR desc
    for i, r in enumerate(sorted(summary_rows, key=lambda x: x["mrr"], reverse=True)):
        row = [
            STRATEGY_LABELS.get(r["config_name"], r["config_name"]),
            r["total_chunks"] or "-",
            round(r["avg_chunk_size"] or 0, 1),
            int(r["median_chunk_size"] or 0),
            int(r["min_chunk_size"] or 0),
            int(r["max_chunk_size"] or 0),
            round(r["ingestion_time_seconds"] or 0, 1),
            round(r["precision_at_5"] or 0, 4),
            round(r["recall_at_5"] or 0, 4),
            round(r["f1_at_5"] or 0, 4),
            round(r["hit_rate_at_5"] or 0, 4),
            round(r["mrr"] or 0, 4),
            round(r["retrieval_latency"] or 0, 1),
        ]
        ws1.append(row)
        for col in range(1, len(headers) + 1):
            c = ws1.cell(row=ws1.max_row, column=col)
            c.fill = ALT_FILLS[i % 2]
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center")

    # Highlight MRR column best value (col 12)
    data_start = 4  # row 4 is first data row
    data_end = data_start + len(summary_rows) - 1
    mrr_vals = [ws1.cell(row=r, column=12).value for r in range(data_start, data_end + 1)]
    best_mrr = max(mrr_vals)
    for r in range(data_start, data_end + 1):
        if ws1.cell(row=r, column=12).value == best_mrr:
            ws1.cell(row=r, column=12).font = Font(bold=True, color="006400")
            ws1.cell(row=r, column=11).font = Font(bold=True, color="006400")

    auto_col_width(ws1)
    ws1.freeze_panes = "B4"

    # ── Sheet 2: Per-Question Detail ─────────────────────────────────────────
    ws2 = wb.create_sheet("Chi tiết câu hỏi")
   
    ws2["A1"] = "CHI TIẾT KẾT QUẢ TỪNG CÂU HỎI – 7 CHIẾN LƯỢC"
    ws2["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws2.merge_cells("A1:G1")
   
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.append([])

    h2 = ["Chiến lược", "Câu hỏi ID", "HitRate", "MRR", "Ref Coverage", "Full Hit", "Latency (ms)"]
    ws2.append(h2)
    make_header_style(ws2, ws2.max_row, len(h2))

    prev_strat = None
    alt_idx = 0
    for r in per_q_rows:
        strat = STRATEGY_LABELS.get(r["config_name"], r["config_name"])
        if strat != prev_strat:
            alt_idx = (alt_idx + 1) % 2
            prev_strat = strat

        hit_val = round(r["hit_rate"] or 0, 2)
        row = [
            strat,
            r["question_id"],
            hit_val,
            round(r["mrr"] or 0, 4),
            round(r["ref_coverage"] or 0, 4),
            round(r["full_hit"] or 0, 2),
            round(r["retrieval_latency"] or 0, 1),
        ]
        ws2.append(row)
        fill = ALT_FILLS[alt_idx]
        for col in range(1, len(h2) + 1):
            c = ws2.cell(row=ws2.max_row, column=col)
            c.fill = fill
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center")
        # Tô đỏ nhạt các ô miss (hit=0)
        if hit_val == 0:
            for col in [3, 4]:
                ws2.cell(row=ws2.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")

    auto_col_width(ws2)
    ws2.freeze_panes = "A4"

    # ── Sheet 3: Câu hỏi bị miss ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Phân tích Miss")
   
    ws3["A1"] = "PHÂN TÍCH CÂU HỎI BỊ TRUY XUẤT SAI (HIT = 0)"
    ws3["A1"].font = Font(bold=True, size=13, color="C00000")
    ws3.merge_cells("A1:H1")
   
    ws3["A1"].alignment = Alignment(horizontal="center")
    ws3.append([])

    h3 = ["QID", "Câu hỏi (trích)", "GT Docs", "GT Pages",
          "Retrieved Pages (top-5)", "Miss ở N chiến lược", "HitRate", "MRR"]
    ws3.append(h3)
    make_header_style(ws3, ws3.max_row, len(h3), fill_hex="C00000")

    for i, m in enumerate(miss_table):
        row = [
            m["question_id"],
            m["question"][:80],
            ", ".join(str(d) for d in m["gt_docs"]),
            ", ".join(str(p) for p in m["gt_pages"]),
            ", ".join(str(p) for p in m["ret_pages"]),
            m["miss_count"],
            round(m["hit_rate"], 2),
            round(m["mrr"], 4),
        ]
        ws3.append(row)
        for col in range(1, len(h3) + 1):
            c = ws3.cell(row=ws3.max_row, column=col)
            c.fill = ALT_FILLS[i % 2]
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left", wrap_text=True)

    auto_col_width(ws3)
    ws3.freeze_panes = "A4"

    # ── Sheet 4: Chunk Statistics ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Chunk Statistics")
   
    ws4["A1"] = "THỐNG KÊ CHUNK – 7 CHIẾN LƯỢC CHUNKING"
    ws4["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws4.merge_cells("A1:G1")
   
    ws4["A1"].alignment = Alignment(horizontal="center")
    ws4.append([])

    h4 = ["Chiến lược", "Tổng chunks", "Avg size (chars)", "Median size", "Min size", "Max size", "Build time (s)"]
    ws4.append(h4)
    make_header_style(ws4, ws4.max_row, len(h4))

    for i, r in enumerate(summary_rows):
        row = [
            STRATEGY_LABELS.get(r["config_name"], r["config_name"]),
            r["total_chunks"] or "-",
            round(r["avg_chunk_size"] or 0, 1),
            int(r["median_chunk_size"] or 0),
            int(r["min_chunk_size"] or 0),
            int(r["max_chunk_size"] or 0),
            round(r["ingestion_time_seconds"] or 0, 1),
        ]
        ws4.append(row)
        for col in range(1, len(h4) + 1):
            c = ws4.cell(row=ws4.max_row, column=col)
            c.fill = ALT_FILLS[i % 2]
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center")

    auto_col_width(ws4)

    # ── Save ─────────────────────────────────────────────────────────────────
   
    out_path = Path(out_path) if out_path else OUT_DIR / "thesis_benchmark_report.xlsx"
    try:
        wb.save(out_path)
    except PermissionError:
        fallback = out_path.with_name(f"{out_path.stem}_restored{out_path.suffix}")
        wb.save(fallback)
        print(f"[WARN] Không ghi được {out_path} (file đang mở). Đã lưu: {fallback}")
        return fallback
    print(f"[OK] Excel saved: {out_path}")
    return out_path

# ─── 5. Charts ───────────────────────────────────────────────────────────────
def make_charts(summary_rows):
    # Sort by strategy order
    strat_map = {r["config_name"]: r for r in summary_rows}
    ordered = [strat_map[s] for s in STRATEGY_ORDER if s in strat_map]
    labels  = [STRATEGY_LABELS[r["config_name"]] for r in ordered]

    metrics = {
        "MRR":          [r["mrr"]           for r in ordered],
        "HitRate@5":    [r["hit_rate_at_5"] for r in ordered],
        "Recall@5":     [r["recall_at_5"]   for r in ordered],
        "Precision@5":  [r["precision_at_5"]for r in ordered],
        "F1@5":         [r["f1_at_5"]       for r in ordered],
    }
    latency = [r["retrieval_latency"] for r in ordered]
    chunks  = [r["total_chunks"]      for r in ordered]

    COLORS = ["#2563EB", "#16A34A", "#D97706", "#DC2626", "#7C3AED", "#0891B2", "#BE185D"]

    # ── Figure 1: Main metrics bar chart ─────────────────────────────────────
    fig, axes = plt.subplots(1, 2, figsize=(18, 7))
    fig.suptitle("So sánh Chiến lược Chunking – Dataset v5.0 (10 câu hỏi)",
                 fontsize=16, fontweight='bold', y=1.01)

    ax = axes[0]
    x  = np.arange(len(labels))
    width = 0.15
    metric_items = list(metrics.items())
    metric_colors = ["#1e40af", "#15803d", "#b45309", "#b91c1c", "#6d28d9"]

    for i, (metric_name, vals) in enumerate(metric_items):
        offset = (i - len(metric_items)//2) * width
        bars = ax.bar(x + offset, vals, width, label=metric_name,
                      color=metric_colors[i], alpha=0.88, edgecolor="white", linewidth=0.5)
        for bar, val in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.004,
                    f"{val:.3f}", ha='center', va='bottom', fontsize=6.5, rotation=90)

    ax.set_title("Metrics Retrieval (Precision/Recall/F1/HR@5/MRR)", fontsize=12, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=18, ha='right', fontsize=10)
    ax.set_ylim(0, 1.15)
    ax.set_ylabel("Score", fontsize=11)
    ax.legend(loc='upper right', fontsize=9)
    ax.yaxis.grid(True, linestyle='--', alpha=0.5)
    ax.set_axisbelow(True)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    # ── Latency chart ─────────────────────────────────────────────────────────
    ax2 = axes[1]
    bar_colors = [COLORS[i % len(COLORS)] for i in range(len(labels))]
    bars2 = ax2.bar(labels, latency, color=bar_colors, alpha=0.88,
                    edgecolor="white", linewidth=0.7)
    for bar, val in zip(bars2, latency):
        ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 30,
                 f"{val:.0f}ms", ha='center', va='bottom', fontsize=9, fontweight='bold')
    ax2.set_title("Latency Trung bình (ms) – Top-5 Retrieval", fontsize=12, fontweight='bold')
    ax2.set_xticklabels(labels, rotation=18, ha='right', fontsize=10)
    ax2.set_ylabel("Latency (ms)", fontsize=11)
    ax2.yaxis.grid(True, linestyle='--', alpha=0.5)
    ax2.set_axisbelow(True)
    ax2.spines['top'].set_visible(False)
    ax2.spines['right'].set_visible(False)

    plt.tight_layout()
   
    path1 = OUT_DIR / "chart_metrics_comparison.png"
    fig.savefig(path1, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"[OK] Chart 1 saved: {path1}")

    # ── Figure 2: Radar chart ─────────────────────────────────────────────────
    radar_metrics = ["MRR", "HitRate@5", "Recall@5", "F1@5", "Precision@5"]
    N = len(radar_metrics)
    angles = [n / float(N) * 2 * np.pi for n in range(N)]
    angles += angles[:1]

    fig2, ax3 = plt.subplots(figsize=(10, 10), subplot_kw=dict(polar=True))
    ax3.set_theta_offset(np.pi / 2)
    ax3.set_theta_direction(-1)
    ax3.set_xticks(angles[:-1])
    ax3.set_xticklabels(radar_metrics, fontsize=11)
    ax3.set_ylim(0, 1)
    ax3.set_yticks([0.2, 0.4, 0.6, 0.8, 1.0])
    ax3.set_yticklabels(["0.2","0.4","0.6","0.8","1.0"], fontsize=8)

    for i, r in enumerate(ordered):
        vals_r = [
            r["mrr"] or 0,
            r["hit_rate_at_5"] or 0,
            r["recall_at_5"] or 0,
            r["f1_at_5"] or 0,
            r["precision_at_5"] or 0,
        ]
        vals_r += vals_r[:1]
        ax3.plot(angles, vals_r, linewidth=2, linestyle='solid',
                 label=STRATEGY_LABELS[r["config_name"]], color=COLORS[i % len(COLORS)])
        ax3.fill(angles, vals_r, alpha=0.08, color=COLORS[i % len(COLORS)])

    ax3.set_title("Radar Chart – So sánh đa chiều các Chiến lược Chunking\n(Dataset v5.0, 10 câu hỏi)",
                  fontsize=13, fontweight='bold', pad=20)
    ax3.legend(loc='upper right', bbox_to_anchor=(1.35, 1.1), fontsize=10)

   
    path2 = OUT_DIR / "chart_radar_comparison.png"
    fig2.savefig(path2, dpi=150, bbox_inches='tight')
    plt.close(fig2)
    print(f"[OK] Chart 2 (Radar) saved: {path2}")

    # ── Figure 3: Miss analysis per strategy ──────────────────────────────────
    return path1, path2

def make_miss_chart(per_strategy):
    strats_ordered = [s for s in STRATEGY_ORDER if s in per_strategy]
    miss_counts    = [len(per_strategy[s]["miss"])    for s in strats_ordered]
    partial_counts = [len(per_strategy[s]["partial"]) for s in strats_ordered]
    hit_counts     = [len(per_strategy[s]["hit"])     for s in strats_ordered]
    xlabels = [STRATEGY_LABELS[s] for s in strats_ordered]

    x = np.arange(len(xlabels))
    fig, ax = plt.subplots(figsize=(13, 6))

    p1 = ax.bar(x, hit_counts,    color="#16A34A", label="Hit (đúng hoàn toàn)", alpha=0.9)
    p2 = ax.bar(x, partial_counts,bottom=hit_counts, color="#F59E0B", label="Partial hit", alpha=0.9)
    p3 = ax.bar(x, miss_counts,
                bottom=[h+p for h,p in zip(hit_counts, partial_counts)],
                color="#EF4444", label="Miss (truy xuất sai hoàn toàn)", alpha=0.9)

    for bar, val in zip(p1, hit_counts):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height()/2,
                    str(val), ha='center', va='center', fontsize=10, color='white', fontweight='bold')
    for bar, val, base in zip(p2, partial_counts, hit_counts):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width()/2, base + val/2,
                    str(val), ha='center', va='center', fontsize=10, color='white', fontweight='bold')
    for bar, val, base1, base2 in zip(p3, miss_counts, hit_counts, partial_counts):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width()/2, base1 + base2 + val/2,
                    str(val), ha='center', va='center', fontsize=10, color='white', fontweight='bold')

    ax.set_title("Phân tích Hit / Partial Hit / Miss theo Chiến lược\n(Dataset v5.0 – 10 câu hỏi)",
                 fontsize=13, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, rotation=15, ha='right', fontsize=11)
    ax.set_ylabel("Số câu hỏi", fontsize=11)
    ax.set_ylim(0, 12)
    ax.legend(fontsize=10, loc='upper right')
    ax.yaxis.grid(True, linestyle='--', alpha=0.4)
    ax.set_axisbelow(True)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()
   
    path3 = OUT_DIR / "chart_miss_analysis.png"
    fig.savefig(path3, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"[OK] Chart 3 (Miss Analysis) saved: {path3}")
    return path3

# ─── 6. Text report: top miss questions ──────────────────────────────────────
def print_miss_analysis(miss_table, per_strategy):
    print("\n" + "="*80)
    print(" PHÂN TÍCH CÂU HỎI BỊ TRUY XUẤT SAI")
    print("="*80)

    print(f"\n Câu hỏi bị MISS ở nhiều chiến lược nhất (top 10):")
    print(f"{'QID':<12} {'Miss/7':>6}  {'GT Page':>8}  Câu hỏi")
    print("-"*80)
    for m in miss_table[:10]:
        print(f"{m['question_id']:<12} {m['miss_count']:>6}/7  " f"{str(m['gt_pages']):>8}  {m['question'][:60]}")

    print()
    for strat in STRATEGY_ORDER:
        if strat not in per_strategy:
            continue
        d = per_strategy[strat]
        total = len(d["miss"]) + len(d["partial"]) + len(d["hit"])
        print(f"\n[{STRATEGY_LABELS[strat]}]  Hit={len(d['hit'])}  Partial={len(d['partial'])}  Miss={len(d['miss'])}  Total={total}")
        if d["miss"]:
            print(f" Câu bị miss ({len(d['miss'])} câu):")
            for m in d["miss"][:5]:
                print(f" - {m['question_id']}: GT_page={m['gt_pages']} | Retrieved={m['ret_pages']} | {m['question'][:60]}")
            if len(d["miss"]) > 5:
                print(f" ... và {len(d['miss'])-5} câu nữa.")

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("="*60)
    print(" THESIS BENCHMARK REPORT GENERATOR")
    print(f" Dataset: {DATASET_VERSION} (60 câu đầy đủ) | Strategies: {len(STRATEGY_ORDER)}")
    print("="*60)

    # 1. Fetch data
    summary_rows = fetch_summary()
    per_q_rows   = fetch_per_question()
    print(f"[OK] Loaded {len(summary_rows)} experiments, {len(per_q_rows)} per-question records")

    # 2. Load & analyse detail JSONs
   
    all_details = load_all_details()
    miss_table, per_strategy = analyse_errors(all_details)
    print(f"[OK] Loaded details: {sum(len(v) for v in all_details.values())} question records")

    # 3. Print text analysis
    print_miss_analysis(miss_table, per_strategy)

    # 4. Charts
    print("\n--- Generating charts ---")
    path1, path2 = make_charts(summary_rows)
   
    path3 = make_miss_chart(per_strategy)

    # 5. Excel
    print("\n--- Generating Excel report ---")
    excel_path = export_excel(summary_rows, per_q_rows, per_strategy, miss_table)

    print("\n" + "="*60)
    print(" XONG. Các file đầu ra:")
    print(f" {path1}")
    print(f" {path2}")
    print(f" {path3}")
    if excel_path:
        print(f" {excel_path}")
    print("="*60)

if __name__ == "__main__":
    main()
